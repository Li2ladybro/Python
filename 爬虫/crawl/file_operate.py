import openpyxl
from openpyxl.styles import Font        # 字体
from openpyxl.styles import Alignment   # 对齐方式
def output_file(property_columns:list,repeat_columns:list,data_list:list)->None:
    """
    need package: openpyxl
    use class: Font && Alignment
    openpyxl.styles.Font        # 字体
    openpyxl.styles.Alignment   # 对齐方式

    :param property_columns: 表头信息
    :param repeat_columns: 重复的列序号(这里的序列号是从0开始计数的)
    :param data_list: 记录
    :return: 传入 表头信息，重复的序列号以及记录则返回文件
    """
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 获取活动的工作表，并命名为Out_sheet
    sheet=workbook.active
    sheet.title='Out_sheet'

    # a.表头信息
    # b.重复的列序号(这里的序列号是从0开始计数的)
    # c.记录
    # property_columns = ['学号', '政治', '英语', '数学', '体育', '姓名', '政治面貌', '性别', '班级']
    # repeat_columns=[[0, 5], [1, 10]]
    # data_list = [['1', None, '90', '89', '88', None, '余家耀', '团员', '男', '2班', None], ['2', None, None, None, None, None, '李天林', '团员', '男', '1班', None], ['3', '98', '89', '98', '87', None, '吕淑芳', '党员', '女', '1班', None], ['4', None, '88', '89', '85', None, '叶文松', '团员', '男', '2班', None], ['5', '98', '96', '96', '74', None, '杨康', '团员', '男', '1班', None], ['6', None, None, None, None, None, '刘金治', '团员', '男', '1班', None], ['7', '95', '84', '82', '92', None, '王浩宇', '团员', '男', '2班', None], ['8', '85', '82', '87', '90', None, '刘忆竺', '团员', '女', '1班', None], ['9', '85', None, None, None, None, '潘欣', '党员', '女', '1班', None], ['10', '85', '93', '68', '89', None, '黄洪军', '团员', '男', '2班', None], ['11', '85', '89', '88', '88', None, '孟楠', '团员', '男', '1班', None], ['12', '83', '75', '86', '87', None, '余曦晨', '党员', '男', '1班', None], ['13', '89', '77', '78', '76', None, '吕玉凤', '团员', '女', '3班', None], ['14', '96', '78', '98', '90', None, '马丽莉', '党员', '女', '1班', None], ['15', '74', '65', '88', '88', None, '龙承铭', '团员', '男', '3班', None], ['16', '96', '89', '80', '87', None, '沈佳佳', '党员', '女', '1班', None], ['17', '66', '70', '80', '89', None, '苏子恒', '团员', '男', '1班', None], ['18', None, None, None, None, None, '苏文迪', '团员', '男', '3班', None], ['19', None, None, None, None, None, '余昊硕', '团员', '男', '1班', None]]

    # 获得要删除的列号
    delete_columns=[]
    for index in range(len(repeat_columns)):
        delete_columns.append((repeat_columns[index][1]))

    # 写入表头信息
    for column,value in enumerate(property_columns):
        cell=sheet.cell(row=1, column=column + 1, value=value)
        # 表头：黑体 14号 加粗
        cell.font = Font(name='SimHei', size=14, bold=True)
        # 设置对齐方式  horizontal：水平方向   vertical：垂直方向
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 将记录写入工作表
    for row,block_data in enumerate(data_list):   # row确定所在行，block_data为行记录
        column=1
        for index,value in enumerate(block_data): # column确定列，value为单元格数据
            # 将数据写入 row+2 行，column 列
            if index not in delete_columns:
                cell=sheet.cell(row=row + 2, column=column, value=value)
                # 正文：宋体 12号 蓝色
                cell.font = Font(name='SimSun', size=12,color='0000ff' )
                # 设置对齐方式  horizontal：水平方向   vertical：垂直方向
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 数字格式
                # '0'：整数格式。
                # '0.00'：保留两位小数的格式。
                # '#,##0.00'：千分位分隔符，保留两位小数的格式。
                # '0%'：百分比格式，显示为整数。
                # '0.00%'：百分比格式，保留两位小数。
                # 'yyyy-mm-dd'：日期格式，显示为年 - 月 - 日。
                cell.number_format = '0'
                column+=1
    # 保存工作簿到文件
    workbook.save('爬取信息汇总Output.xlsx')
    # 关闭文件
    workbook.close()
    return

import os
def remove_excel_file(file_path:str)->None:
    """
    need package: os
    删除 Excel文件
    :param file_path:  指定要删除的 Excel 文件路径
    :return: None
    """

    # 检查文件是否存在
    if os.path.exists(file_path):
        # 删除文件
        os.remove(file_path)
        print(f"文件 {file_path} 已删除")
    else:
        print(f"文件 {file_path} 不存在")
    return

from pathlib import Path
def rename_file(old_file_path:str,new_file_path:str)->None:
    """
    重命名文件
    need package: pathlib
    :param old_file_path:   原文件地址
    :param new_file_path:   新文件地址
    :return: None
    """
    # 指定旧文件路径和新文件路径
    old_file_path = Path(old_file_path)
    new_file_path = Path(new_file_path)

    # 检查旧文件是否存在
    if old_file_path.exists():
        # 重命名文件
        old_file_path.rename(new_file_path)
        print(f"文件已从 {old_file_path} 重命名为 {new_file_path}")
    else:
        print(f"文件 {old_file_path} 不存在")

    return